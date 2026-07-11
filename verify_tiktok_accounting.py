import openpyxl

path = r"C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
headers = [str(cell.value).strip() for cell in ws[1]]

def safe_float(val):
    if val is None or val == '': return 0.0
    try: return float(str(val).replace(',', ''))
    except: return 0.0

idx_total_rev = headers.index('Total Pendapatan')
idx_total_biaya = headers.index('Total Biaya')
idx_settle = headers.index('Jumlah penyelesaian pembayaran')
idx_voucher_seller = headers.index('Diskon voucher yang ditanggung penjual')
idx_voucher_platform = headers.index('Diskon voucher yang ditanggung platform')
idx_buyer_pay = headers.index('Pembayaran oleh pembeli')

print("Row-by-row mathematical check of TikTok income:")
print("Let's see if we can find a formula that holds for all rows with voucher seller.")
count = 0
for r in range(2, ws.max_row + 1):
    voucher_seller = safe_float(ws.cell(row=r, column=idx_voucher_seller+1).value)
    if voucher_seller != 0:
        total_rev = safe_float(ws.cell(row=r, column=idx_total_rev+1).value)
        total_biaya = safe_float(ws.cell(row=r, column=idx_total_biaya+1).value)
        settle = safe_float(ws.cell(row=r, column=idx_settle+1).value)
        buyer_pay = safe_float(ws.cell(row=r, column=idx_buyer_pay+1).value)
        voucher_platform = safe_float(ws.cell(row=r, column=idx_voucher_platform+1).value)
        
        # Let's print out the values
        print(f"Row {r}:")
        print(f"  - Buyer Pays: {buyer_pay}")
        print(f"  - Total Pendapatan: {total_rev}")
        print(f"  - Voucher Seller: {voucher_seller}")
        print(f"  - Voucher Platform: {voucher_platform}")
        print(f"  - Total Biaya: {total_biaya}")
        print(f"  - Settlement: {settle}")
        
        # Let's check:
        # Is there any hidden fee or adjustment?
        # Let's check the sum of all columns for this row to see if everything balances to 0 or Settlement
        # The sum of all columns that are part of cash flow:
        # Settlement = (Buyer Pays) - (Total Biaya? No, Total Biaya is platform fees)
        # Wait, does the buyer pay the money to the platform first, and then the platform pays the seller?
        # Yes! The platform settles: Settlement = Total Pendapatan + Total Biaya.
        # But if the seller voucher is -11,070, and it is NOT in Total Pendapatan and NOT in Total Biaya,
        # then does that mean the seller received the full Total Pendapatan (which is BEFORE voucher)?
        # Let's check: Yes! Total Pendapatan is 111,930. The seller received 111,930 - 32,397 = 79,533.
        # So the seller was NOT charged for the voucher in this transaction!
        # Wait! Is it possible that the voucher is charged in a separate penyesuaian (adjustment) row?
        # Let's check if there are any Penyesuaian rows in the spreadsheet!
        count += 1
        if count >= 5:
            break
wb.close()
