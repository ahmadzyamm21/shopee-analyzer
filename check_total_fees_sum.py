import openpyxl

path = r"C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

headers = [str(cell.value).strip() for cell in ws[1]]

def safe_float(val):
    if val is None or val == '': return 0.0
    try: return float(str(val).replace(',', ''))
    except: return 0.0

sub_fees = [
    'Biaya komisi platform', 
    'Komisi Afiliasi', 
    'Komisi mitra afiliasi', 
    'Komisi Iklan Toko afiliasi', 
    'Komisi dinamis', 
    'Biaya layanan logistik', 
    'Ongkir', 
    'Biaya layanan cashback bonus', 
    'Biaya pemrosesan pesanan'
]

total_biaya_sum = 0.0
sub_fees_sums = {name: 0.0 for name in sub_fees}

for r in range(2, ws.max_row + 1):
    t_type = str(ws.cell(row=r, column=headers.index('Jenis transaksi')+1).value).strip()
    if t_type != 'Pesanan': continue
    
    total_biaya_sum += safe_float(ws.cell(row=r, column=headers.index('Total Biaya')+1).value)
    for name in sub_fees:
        if name in headers:
            sub_fees_sums[name] += safe_float(ws.cell(row=r, column=headers.index(name)+1).value)

print(f"Total Biaya column sum: {total_biaya_sum:,.2f}")
print("Sum of sub-fees:")
sum_sub_fees_total = 0.0
for name, val in sub_fees_sums.items():
    print(f" - {name}: {val:,.2f}")
    sum_sub_fees_total += val

print(f"Total sum of sub-fees: {sum_sub_fees_total:,.2f}")
wb.close()
