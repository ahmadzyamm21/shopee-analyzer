import openpyxl

path = r"C:\Users\ahmad\Downloads\Rincian_Biaya_Shopee_Per_Kategori.xlsx"
wb = openpyxl.load_workbook(path, read_only=True)
print("Sheet Names in Rincian_Biaya_Shopee_Per_Kategori:", wb.sheetnames)
ws = wb.active
for r in range(1, 15):
    row_vals = [cell.value for cell in ws[r]]
    if any(row_vals):
        print(f"Row {r}: {row_vals}")
wb.close()
