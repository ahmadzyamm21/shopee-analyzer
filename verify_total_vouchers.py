import openpyxl

path = r"C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active
headers = [str(cell.value).strip() for cell in ws[1]]

def safe_float(val):
    if val is None or val == '': return 0.0
    try: return float(str(val).replace(',', ''))
    except: return 0.0

idx_voucher_seller = headers.index('Diskon voucher yang ditanggung penjual')
idx_voucher_platform = headers.index('Diskon voucher yang ditanggung platform')
idx_diskon_platform = headers.index('Diskon platform')

sum_voucher_seller = 0.0
sum_voucher_platform = 0.0
sum_diskon_platform = 0.0

for r in range(2, ws.max_row + 1):
    sum_voucher_seller += safe_float(ws.cell(row=r, column=idx_voucher_seller+1).value)
    sum_voucher_platform += safe_float(ws.cell(row=r, column=idx_voucher_platform+1).value)
    sum_diskon_platform += safe_float(ws.cell(row=r, column=idx_diskon_platform+1).value)

print(f"Total Diskon voucher yang ditanggung penjual: {sum_voucher_seller:,.2f}")
print(f"Total Diskon voucher yang ditanggung platform: {sum_voucher_platform:,.2f}")
print(f"Total Diskon platform: {sum_diskon_platform:,.2f}")

wb.close()
