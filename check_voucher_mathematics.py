import openpyxl

path = r"C:\Users\ahmad\Downloads\testing maret\income maret 2026.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb.active

headers = [str(cell.value).strip() for cell in ws[1]]

def safe_float(val):
    if val is None or val == '': return 0.0
    try: return float(str(val).replace(',', ''))
    except: return 0.0

idx_total_rev = headers.index('Total Pendapatan')
idx_total_biaya = headers.index('Total Biaya')
idx_settle = headers.index('Jumlah penyelesaian pembayaran')
idx_diskon_voucher_seller = headers.index('Diskon voucher yang ditanggung penjual')
idx_buyer_pay = headers.index('Pembayaran oleh pembeli')

print("Analyzing rows with seller voucher:")
count = 0
for r in range(2, ws.max_row + 1):
    total_rev = safe_float(ws.cell(row=r, column=idx_total_rev+1).value)
    total_biaya = safe_float(ws.cell(row=r, column=idx_total_biaya+1).value)
    settle = safe_float(ws.cell(row=r, column=idx_settle+1).value)
    diskon_voucher_seller = safe_float(ws.cell(row=r, column=idx_diskon_voucher_seller+1).value)
    buyer_pay = safe_float(ws.cell(row=r, column=idx_buyer_pay+1).value)
    
    if diskon_voucher_seller != 0:
        print(f"\nRow {r}:")
        print(f"  - Buyer Pays: {buyer_pay}")
        print(f"  - Total Pendapatan (Omset): {total_rev}")
        print(f"  - Diskon Voucher Seller (Voucher Toko): {diskon_voucher_seller}")
        print(f"  - Total Biaya: {total_biaya}")
        print(f"  - Settlement: {settle}")
        
        # Let's see: how is Buyer Pays related to Total Pendapatan + Ongkir etc.
        # Usually, Total Pendapatan is what the buyer pays for the product, excluding shipping.
        # If the buyer uses a seller voucher, does the buyer pay less?
        # Yes, buyer pays 105,596, but Total Pendapatan is 111,930.
        # Why is Total Pendapatan higher than buyer pays?
        # Because Total Pendapatan = 111,930. If voucher of 11,070 is applied,
        # then the actual product price paid by buyer is 111,930 - 11,070 = 100,860.
        # If buyer pays 105,596, that means buyer paid 100,860 + shipping (which is 4,736?).
        # If the seller receives 79,533, and HPP is HPP.
        # Is the seller voucher ALREADY subtracted from Total Pendapatan?
        # NO! Total Pendapatan is 111,930. If it was subtracted, it would be 100,860!
        # This means the seller's settlement of 79,533 is calculated as 111,930 - 32,397.
        # That means the seller received 79,533.
        # But wait! If the seller received 79,533, did the seller actually get charged for the voucher?
        # If the seller received 79,533, they did NOT get charged the 11,070 voucher!
        # Because if they got charged, the settlement would be 79,533 - 11,070 = 68,463!
        # Wait, is there a voucher adjustment or does TikTok reimburse it?
        # Let's print all other voucher columns for Row 233 again:
        #   - 71: Diskon platform = 7,380.0
        #   - 73: Diskon voucher yang ditanggung platform = 7,380.0
        # Wait! Diskon platform is positive 7,380!
        # Let's print all columns for Row 233 that contain "voucher" or "diskon" or "subsidi"
        count += 1
        if count >= 3:
            break

wb.close()
