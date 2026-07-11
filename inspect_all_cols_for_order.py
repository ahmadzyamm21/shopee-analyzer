import openpyxl

order_path = r"C:\Users\ahmad\Downloads\testing maret\pesanan maret 2026.xlsx"
wb = openpyxl.load_workbook(order_path, data_only=True)
ws = wb['OrderSKUList']

headers = [str(cell.value).strip() for cell in ws[1]]
oid_col = headers.index('Order ID')

print(f"Total rows in sheet: {ws.max_row}")
print("\nRow details for Order ID 583293734764053945:")
for r in range(2, ws.max_row + 1):
    val = str(ws.cell(row=r, column=oid_col+1).value).strip()
    if val == '583293734764053945':
        print(f"\nRow {r}:")
        for idx, h in enumerate(headers):
            cell_val = ws.cell(row=r, column=idx+1).value
            print(f"  - [{idx}] {h}: {cell_val}")
        break
wb.close()
